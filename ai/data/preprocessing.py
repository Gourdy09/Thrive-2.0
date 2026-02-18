"""
1. Reads 593 meals + 15397 glucose readings
2. Assigns medication periods based on date
3. Creates sequences
4. Saves train/val/test JSON
"""

import numpy as np 
from typing import Dict, List, Tuple, Any
from datetime import datetime as dt, time as time_type
from pathlib import Path 
import json 
import openpyxl


MEDICATION_PERIODS = {
    'baseline': {'start': dt(2025, 6, 12), 'end': dt(2025, 6, 26)},
    'pioglitazone_45': {'start': dt(2025, 6, 27), 'end': dt(2025, 9, 16)},
    'metformin_500': {'start': dt(2025, 9, 17), 'end': dt(2025, 10, 23)},
    'metformin_500_er': {'start': dt(2025, 10, 24), 'end': dt(2026, 1, 18)},
}

def get_med_period(meal_date: dt) -> str:
    for period_name, period_info in MEDICATION_PERIODS.items():
        if period_info['start'] <= meal_date <= period_info['end']:
            return period_name
    return 'unknown'

def read_excel_data(excel_path: str) -> Tuple[List[Dict], List[Dict]]:
    print(f"Reading Excel file...")
    
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active
    
    meals_data = []
    glucose_readings = []
    current_date = None
    
    for row_idx in range(3, ws.max_row + 1):
        date_cell = ws.cell(row_idx, 1).value
        time_cell = ws.cell(row_idx, 2).value
        carbs_cell = ws.cell(row_idx, 3).value
        glucose_cell = ws.cell(row_idx, 5).value
        fasting_cell = ws.cell(row_idx, 6).value
        
        if isinstance(date_cell, dt):
            current_date = date_cell
        
        if not current_date:
            continue
        
        if not isinstance(time_cell, time_type):
            continue
        
        hour = time_cell.hour + time_cell.minute / 60
        
        if carbs_cell and isinstance(carbs_cell, (int, float)) and carbs_cell > 0:
            meals_data.append({
                'date': current_date,
                'hour': hour,
                'carbs': float(carbs_cell),
                'is_liquid': False,
                'fiber_ratio': 0.15,
                'fatprotein': 0.2,
                'activity_level': 0.0,
                'medication_period': get_med_period(current_date)
            })
        
        if glucose_cell and isinstance(glucose_cell, (int, float)):
            glucose_readings.append({
                'date': current_date,
                'hour': hour,
                'glucose': float(glucose_cell),
                'is_fasting': bool(fasting_cell)
            })
    
    print(f"✅ Read {len(meals_data)} meals, {len(glucose_readings)} glucose readings")
    return meals_data, glucose_readings

def create_sequences(meals: List[Dict], glucose: List[Dict]) -> List[Tuple[Dict, Dict]]:
    print("Creating sequences...")
    sequences = []
    
    for meal in meals:
        meal_date = meal['date']
        meal_hour = meal['hour']
        
        day_glucose = [g for g in glucose if g['date'] == meal_date]
        if not day_glucose:
            continue
        
        day_glucose.sort(key=lambda x: x['hour'])
        
        baseline = [g for g in day_glucose if g['hour'] < meal_hour]
        if baseline:
            baseline_glucose = baseline[-1]['glucose']
        else:
            baseline_glucose = day_glucose[0]['glucose']
        
        post_meal = [g for g in day_glucose if meal_hour <= g['hour'] <= meal_hour + 3]
        if len(post_meal) < 2:
            continue
        
        times = [max(0, (g['hour'] - meal_hour) * 60) for g in post_meal]
        values = [g['glucose'] for g in post_meal]
        
        sequences.append((meal, {
            'baseline_glucose': baseline_glucose,
            'glucose_times': times,
            'glucose_values': values
        }))
    
    print(f" Created {len(sequences)} sequences")
    return sequences

def split_data(sequences: List, train_pct=0.70, val_pct=0.15, seed=42):
    np.random.seed(seed)
    n = len(sequences)
    indices = np.random.permutation(n)
    
    train_end = int(n * train_pct)
    val_end = train_end + int(n * val_pct)
    
    train = [sequences[i] for i in indices[:train_end]]
    val = [sequences[i] for i in indices[train_end:val_end]]
    test = [sequences[i] for i in indices[val_end:]]
    
    print(f" Split: Train={len(train)} Val={len(val)} Test={len(test)}")
    return train, val, test

def save_data(train, val, test, output_dir="./data/processed"):
    Path(output_dir).mkdir(parents=True, exist_ok=True)
    
    for name, data in [("train", train), ("val", val), ("test", test)]:
        dataset = {
            'meal_features': [
                {
                    'carbs': m['carbs'],
                    'hour': m['hour'],
                    'fiber_ratio': m['fiber_ratio'],
                    'is_liquid': m['is_liquid'],
                    'fatprotein': m['fatprotein'],
                    'activity_level': m['activity_level'],
                    'medication_period': m['medication_period'],
                    'date': m['date'].isoformat()
                }
                for m, _ in data
            ],
            'glucose_sequences': [seq for _, seq in data]
        }
        
        with open(f"{output_dir}/{name}.json", 'w') as f:
            json.dump(dataset, f, indent=2)
        print(f" Saved {name}.json ({len(dataset['meal_features'])} examples)")
    
    return train, val, test

def create_metadata(train, val, test, output_dir="./data/processed"):
    all_meals = [m for m, _ in train + val + test]
    all_glucose_seqs = [seq for _, seq in train + val + test]
    
    all_glucose = []
    for seq in all_glucose_seqs:
        all_glucose.extend(seq['glucose_values'])
    
    all_carbs = [m['carbs'] for m in all_meals]
    
    med_dist = {}
    for m in all_meals:
        med = m['medication_period']
        med_dist[med] = med_dist.get(med, 0) + 1
    
    metadata = {
        'num_sequences': len(all_meals),
        'num_glucose_readings': len(all_glucose),
        'train_split': len([m for m, _ in train]),
        'val_split': len([m for m, _ in val]),
        'test_split': len([m for m, _ in test]),
        'glucose_stats': {
            'min': float(np.min(all_glucose)),
            'max': float(np.max(all_glucose)),
            'mean': float(np.mean(all_glucose)),
            'std': float(np.std(all_glucose)),
        },
        'carb_stats': {
            'min': float(np.min(all_carbs)),
            'max': float(np.max(all_carbs)),
            'mean': float(np.mean(all_carbs)),
            'std': float(np.std(all_carbs)),
        },
        'medication_distribution': med_dist,
    }
    
    with open(f"{output_dir}/metadata.json", 'w') as f:
        json.dump(metadata, f, indent=2)
    
    print(f"\n Saved metadata.json")
    print(f" Medication Distribution:")
    for med, count in med_dist.items():
        print(f"   {med}: {count}")

def load_processed_data(data_dir: str = "./data/processed"):
    data_dir = Path(data_dir)
    
    with open(data_dir / "train.json") as f:
        train_data = json.load(f)
    
    with open(data_dir / "val.json") as f:
        val_data = json.load(f)
    
    with open(data_dir / "test.json") as f:
        test_data = json.load(f)
    
    return train_data, val_data, test_data

def main():
    print("=" * 70)
    print("GLUCOSE DATA PREPROCESSING")
    print("=" * 70)
    
    excel_path = "./Dexcom Stelo blood glucose readings - food carbs - exercise - 1 18 2026.xlsx"
    
    if not Path(excel_path).exists():
        print(f"❌ File not found")
        return
    
    meals, glucose = read_excel_data(excel_path)
    
    if not meals or not glucose:
        print("❌ No data read")
        return
    
    sequences = create_sequences(meals, glucose)
    
    if not sequences:
        print("❌ No sequences created")
        return
    
    train, val, test = split_data(sequences)
    train, val, test = save_data(train, val, test)
    create_metadata(train, val, test)
    
    print("\n" + "=" * 70)
    print("✅ PREPROCESSING COMPLETE!")
    print("=" * 70)

if __name__ == "__main__":
    main()