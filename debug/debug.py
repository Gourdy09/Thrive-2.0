"""
DEBUG Preprocessing - Shows what's being read from Excel
"""

import numpy as np 
from typing import Dict, List, Tuple, Any
from datetime import datetime as dt, time as time_type
from pathlib import Path 
import json 
import openpyxl


def read_excel_data_DEBUG(excel_path: str):
    """Debug version that shows exactly what's happening"""
    print(f"Reading Excel file: {excel_path}")
    
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active
    
    print(f"Sheet: {ws.title}")
    print(f"Max rows: {ws.max_row}")
    
    meals_data = []
    glucose_readings = []
    current_date = None
    
    print(f"\nProcessing rows 3-{ws.max_row}...")
    print("=" * 100)
    
    # Check ALL rows
    for row_idx in range(3, ws.max_row + 1):
        date_cell = ws.cell(row_idx, 1).value
        time_cell = ws.cell(row_idx, 2).value
        carbs_cell = ws.cell(row_idx, 3).value
        glucose_cell = ws.cell(row_idx, 5).value
        fasting_cell = ws.cell(row_idx, 6).value
        
        print(f"\nRow {row_idx}:")
        print(f"  A (Date): {date_cell} | Type: {type(date_cell).__name__}")
        print(f"  B (Time): {time_cell} | Type: {type(time_cell).__name__}")
        print(f"  C (Carbs): {carbs_cell} | Type: {type(carbs_cell).__name__ if carbs_cell else 'None'}")
        print(f"  E (Glucose): {glucose_cell} | Type: {type(glucose_cell).__name__ if glucose_cell else 'None'}")
        
        # Process date
        if date_cell:
            if isinstance(date_cell, dt):
                current_date = date_cell
                print(f"  ✅ Updated current_date to {current_date}")
            elif isinstance(date_cell, str):
                print(f"  String in date column: '{date_cell}' (len={len(date_cell)})")
        
        if not current_date:
            print(f"  ⚠️ No current_date yet, skipping row")
            continue
        
        # Process time
        if time_cell:
            try:
                if isinstance(time_cell, time_type):
                    hour = time_cell.hour + time_cell.minute / 60
                    print(f"  ✅ Parsed time: {time_cell} → hour={hour:.2f}")
                else:
                    print(f"  ❌ Time cell not time object, skipping")
                    continue
            except Exception as e:
                print(f"  ❌ Error parsing time: {e}")
                continue
        else:
            print(f"  ⚠️ No time in row, skipping")
            continue
        
        # Process carbs
        if carbs_cell and isinstance(carbs_cell, (int, float)) and carbs_cell > 0:
            print(f"  ✅ MEAL FOUND: {carbs_cell}g carbs")
            meals_data.append({
                'date': current_date,
                'hour': hour,
                'carbs': carbs_cell
            })
        
        # Process glucose
        if glucose_cell and isinstance(glucose_cell, (int, float)):
            print(f"  ✅ GLUCOSE: {glucose_cell} mg/dL")
            glucose_readings.append({
                'date': current_date,
                'hour': hour,
                'glucose': glucose_cell
            })
    
    print("\n" + "=" * 100)
    print(f"\n📊 RESULTS:")
    print(f"   Meals found: {len(meals_data)}")
    print(f"   Glucose readings found: {len(glucose_readings)}")
    
    return meals_data, glucose_readings


if __name__ == "__main__":
    excel_path = "./Dexcom Stelo blood glucose readings - food carbs - exercise - 1 18 2026.xlsx"
    
    if not Path(excel_path).exists():
        print(f"❌ File not found: {excel_path}")
        exit()
    
    meals, glucose = read_excel_data_DEBUG(excel_path)